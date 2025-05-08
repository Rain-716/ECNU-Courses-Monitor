# -*- coding: utf-8 -*-
import os
import re
import json
import time
import pandas
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl.utils import get_column_letter

# 用户配置
USERNAME = '你的学号'
PASSWORD = '你的公共数据库密码'
LOGIN_URL = 'https://applicationnewjw.ecnu.edu.cn/eams/stdElectCourse.action'
TARGET_URL = 'https://applicationnewjw.ecnu.edu.cn/eams/stdElectCourse!queryStdCount.action?profileId=6022'
JS_FILE_COUNTS = 'stdElectCourse!queryStdCount.js'
JS_FILE_LESSONS = 'stdElectCourse!data.js'
OUTPUT_EXCEL = 'lesson_overview_with_counts.xlsx'
INTERVAL_SECONDS = 60  # 循环时间

# 提取 JS 文件中的 JSON 数据

def load_js_json(file_path):
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()
    content = re.sub(r'<.*?>', '', content)
    content = re.sub(r'/\*.*?\*/', '', content, flags=re.S)
    start = content.find('=') + 1
    semicolon_pos = content.rfind(';')
    end = semicolon_pos if semicolon_pos > start else len(content)
    data_str = content[start:end]
    data_str = re.sub(r'([{,]\s*)([a-zA-Z_]\w*)(\s*:)', r'\1"\2"\3', data_str)
    data_str = re.sub(r"'([^']*)'", r'"\1"', data_str)
    return data_str

# 主流程

def run_task():
    # Selenium 登录并获取选课人数 JS 文件
    driver = webdriver.Chrome()
    try:
        driver.get(LOGIN_URL)
        wait = WebDriverWait(driver, 10)
        username_input = wait.until(EC.presence_of_element_located((By.NAME, 'username')))
        password_input = driver.find_element(By.CSS_SELECTOR, "#normalLoginForm input[type='password']")
        username_input.send_keys(USERNAME)
        password_input.send_keys(PASSWORD)

        # captcha_text = input("请输入验证码: ").strip()
        # captcha_input = driver.find_element(By.CSS_SELECTOR, '#normalLoginForm app-verification input')
        # captcha_input.send_keys(captcha_text)
        # driver.find_element(By.ID, 'submitBtn').click()

        enter_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.LINK_TEXT, '进入选课>>>>'))
        )
        enter_button.click()
        driver.get(TARGET_URL)
        time.sleep(1)
        driver.get(TARGET_URL)
        page_content = driver.page_source

        with open(JS_FILE_COUNTS, 'w', encoding='utf-8') as f:
            f.write(page_content)
        print(f"页面内容已保存到 {JS_FILE_COUNTS}")
    finally:
        driver.quit()

    # 解析 lessons 和 counts JS
    json_lessons = load_js_json(JS_FILE_LESSONS)
    lessons = json.loads(json_lessons)

    json_counts = load_js_json(JS_FILE_COUNTS)
    counts_dict = json.loads(json_counts)

    # 构建 DataFrame 并合并
    df = pandas.DataFrame(lessons)
    remove_cols = [
        'arrangeInfo', 'expLessonGroups',
        'code', 'courseId', 'courseTypeId', 'courseTypeCode',
        'textbooks', 'campusCode', 'remark'
    ]
    df.drop(columns=[c for c in remove_cols if c in df.columns], inplace=True)

    counts_df = pandas.DataFrame.from_dict(counts_dict, orient='index')
    counts_df.index.name = 'id'
    counts_df.reset_index(inplace=True)
    counts_df.rename(columns={'sc': 'current_count', 'lc': 'limit_count'}, inplace=True)
    counts_df['id'] = counts_df['id'].astype(df['id'].dtype)

    df = df.merge(counts_df, on='id', how='left')
    df['remain_count'] = df['limit_count'] - df['current_count']

    # 导出到 Excel
    with pandas.ExcelWriter(OUTPUT_EXCEL, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='课程概览')
        workbook = writer.book
        worksheet = writer.sheets['课程概览']
        for idx, col in enumerate(df.columns, 1):
            max_length = len(col)
            for cell in worksheet[get_column_letter(idx)]:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            worksheet.column_dimensions[get_column_letter(idx)].width = max_length + 2
    print(f"已成功将课程信息及人数数据导出到 {OUTPUT_EXCEL}")

if __name__ == '__main__':
    while True:
        run_task()
        time.sleep(INTERVAL_SECONDS)
